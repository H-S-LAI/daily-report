import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import io

# --- 1. 資料清洗 ---
def load_data(file):
    try:
        if file.name.endswith('.xlsx'):
            df = pd.read_excel(file, engine='openpyxl')
        else:
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except:
                df = pd.read_csv(file, encoding='cp950')
        
        rename_map = {
            '店別': '區域', '班別營業|店名': '店名', '班別營業|班別': '班別',
            '班別營業|日期': '日期', '班別營業|值班者': '值班者',
            '檳榔銷售|金額': '檳榔', '營業金額|實收金額': '實收', '營業金額|結帳差額': '帳差'
        }
        
        actual_rename = {}
        for target_key, new_name in rename_map.items():
            for col in df.columns:
                if target_key in col:
                    actual_rename[col] = new_name
        
        df = df[list(actual_rename.keys())].rename(columns=actual_rename)

        for col in ['檳榔', '實收', '帳差']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        
        # [區域歸類] 日紅店必須算在彰化
        if '區域' in df.columns:
            df['區域'] = df['區域'].astype(str).str.strip()
            df.loc[df['區域'].str.contains('日紅|彰化'), '區域'] = '彰化'
            df.loc[df['區域'].str.contains('台中'), '區域'] = '台中'
        
        # 帳差加負號
        if '帳差' in df.columns:
            df['帳差'] = df['帳差'] * -1
        
        report_date = pd.to_datetime(df.iloc[0]['日期']) if not df.empty else datetime.now()
        return df, report_date
    except Exception as e:
        st.error(f"讀取資料失敗：{e}")
        return None, None

# --- 2. 讀取昨日累計 (含公式數值) ---
def get_cumulative(file_content, current_date):
    if current_date.day == 1: return 0, 0, 0
    try:
        # data_only=True 確保讀到公式計算後的數字
        temp_wb = load_workbook(io.BytesIO(file_content), data_only=True)
        names = temp_wb.sheetnames
        prev_name = (current_date - timedelta(days=1)).strftime("%m-%d")
        ws = temp_wb[prev_name] if prev_name in names else temp_wb[names[-1]]
        
        t, ch, tc = 0, 0, 0
        for row in ws.iter_rows(min_col=12, max_col=15):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = ws.cell(row=cell.row, column=16).value or 0
                    if "營業總金額" in cell.value: t = float(val)
                    elif "彰化檳榔金額" in cell.value: ch = float(val)
                    elif "台中檳榔金額" in cell.value: tc = float(val)
        return t, ch, tc
    except Exception as e:
        return 0, 0, 0

# --- 3. 網頁介面 ---
st.set_page_config(page_title="直營店日報產生器 V16", layout="wide")
st.title("🍹 直營店日報自動化系統 V16")

f1 = st.file_uploader("1. 上傳當日系統原始檔", type=['csv', 'xlsx'])
f2 = st.file_uploader("2. 上傳目前的月累計 Excel (非 1 號必傳)", type=['xlsx'])

if st.button("🚀 生成報表"):
    if f1:
        df, report_date = load_data(f1)
        if df is None or df.empty:
            st.error("原始檔內容為空。")
            st.stop()

        if f2:
            f2_content = f2.getvalue()
            p_t, p_ch, p_tc = get_cumulative(f2_content, report_date)
            wb = load_workbook(io.BytesIO(f2_content))
        else:
            p_t, p_ch, p_tc = 0, 0, 0
            wb = Workbook()
            if 'Sheet' in wb.sheetnames: del wb['Sheet']
        
        sn = report_date.strftime("%m-%d")
        if sn in wb.sheetnames: del wb[sn]
        ws = wb.create_sheet(sn)

        # --- 樣式定義 ---
        # 邊框定義
        thin_side = Side('thin')
        blue_side = Side('thin', color='0000FF') # 藍色邊框
        
        border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        # 底部藍線樣式 (用於店鋪分隔)
        border_blue_bottom = Border(left=thin_side, right=thin_side, top=thin_side, bottom=blue_side)

        align_c = Alignment('center', 'center', wrap_text=True)
        align_r = Alignment('right', 'center', wrap_text=True)
        align_l_top = Alignment('left', 'top', wrap_text=True)
        
        # 字體設定 (依需求調整)
        font_title = Font('微軟正黑體', 16, bold=True)      # 標題 16
        font_header = Font('微軟正黑體', 12, bold=False)     # 表頭 12 (預設) - 雖然你沒特別說表頭要改，但配合整體我設12
        font_n = Font('微軟正黑體', 14)                    # 一般數字 14
        font_b = Font('微軟正黑體', 14, bold=True)         # 粗體數字 14
        font_red = Font('微軟正黑體', 14, color="FF0000", bold=True)
        font_blue = Font('微軟正黑體', 14, color="0000FF", bold=True)
        font_green = Font('微軟正黑體', 14, color="008000", bold=True)
        font_panel = Font('微軟正黑體', 14, bold=True)     # Panel B 14
        
        fill_blue = PatternFill('solid', fgColor="D9E1F2")

        col_ws = {'A':12,'B':6,'C':8,'D':9,'E':9,'F':6,'G':9,'H':6,'I':6,'J':6,'K':2,'L':12,'M':6,'N':8,'O':9,'P':9,'Q':6,'R':9,'S':6,'T':6,'U':6}
        for k, v in col_ws.items(): ws.column_dimensions[k].width = v

        # 標題列 (字級 16)
        tw_year = report_date.year - 1911
        date_str = f" {tw_year}年{report_date.month}月{report_date.day}日"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws['A1']=f"{date_str} 直營店營收報表 (彰化區)"; ws['A1'].font=font_title; ws['A1'].alignment=align_c
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=21)
        ws['L1']=f"{date_str} 直營店營收報表 (台中區)"; ws['L1'].font=font_title; ws['L1'].alignment=align_c

        # 表頭 (修正檳榔換行)
        headers = ['店名', '班別', '值班者', '檳榔\n金額', '實收\n金額', '帳差', '合計', '收款', '實差', '現金\n合計']
        for i, h in enumerate(headers):
            for sc in [1, 12]:
                c = ws.cell(row=2, column=sc+i, value=h)
                c.border=border_all_thin; c.alignment=align_c; c.fill=fill_blue
                c.font = font_header # 表頭字體
        ws.row_dimensions[2].height = 30 # 表頭高一點以容納換行

        # --- 繪製函數 ---
        def render_store(df_s, r, cs):
            if df_s.empty: return r
            rows = len(df_s)
            
            # 寫入內容
            for i in range(rows):
                curr = r + i
                d = df_s.iloc[i]
                
                # 設定列高 22
                ws.row_dimensions[curr].height = 22
                
                # 判斷是否為該店最後一行 (決定格線顏色)
                current_border = border_blue_bottom if i == rows - 1 else border_all_thin

                ws.cell(curr, cs+1, d['班別']).alignment=align_c; ws.cell(curr, cs+1).font=font_n; ws.cell(curr, cs+1).border=current_border
                ws.cell(curr, cs+2, d['值班者']).alignment=align_c; ws.cell(curr, cs+2).font=font_n; ws.cell(curr, cs+2).border=current_border
                
                c_b = ws.cell(curr, cs+3, d['檳榔']); c_b.number_format='#,##0'; c_b.font=font_n; c_b.border=current_border
                c_r = ws.cell(curr, cs+4, d['實收']); c_r.number_format='#,##0'; c_r.font=font_n; c_r.border=current_border
                
                dv = d['帳差']; cd = ws.cell(curr, cs+5, dv); cd.number_format='#,##0'; cd.alignment=align_c; cd.border=current_border
                cd.font = font_red if dv<0 else (font_blue if dv>0 else font_n)
                
                # 後面空白格也要畫線
                for x in range(6, 10): ws.cell(curr, cs+x).border=current_border

            # 合併儲存格 (店名、合計) - 這些跨列的儲存格，邊框要特別處理
            # 店名
            ws.merge_cells(start_row=r, start_column=cs, end_row=r+rows-1, end_column=cs)
            c_name = ws.cell(r, cs, df_s.iloc[0]['店名'])
            c_name.font=font_b; c_name.alignment=align_c
            # 補邊框 (OpenPyXL合併單元格邊框需對左上角設定，但底部邊框需對所有底層單元格設定)
            # 這裡簡化處理：對合併區域的四個角或邊緣設定
            for i in range(rows):
                b_style = border_blue_bottom if i == rows - 1 else border_all_thin
                ws.cell(r+i, cs).border = b_style

            # 合計
            ws.merge_cells(start_row=r, start_column=cs+6, end_row=r+rows-1, end_column=cs+6)
            c_tot = ws.cell(r, cs+6, df_s['實收'].sum())
            c_tot.font=font_b; c_tot.alignment=align_c; c_tot.number_format='#,##0'
            for i in range(rows):
                b_style = border_blue_bottom if i == rows - 1 else border_all_thin
                ws.cell(r+i, cs+6).border = b_style
            
            # 現金合計(空白)
            ws.merge_cells(start_row=r, start_column=cs+9, end_row=r+rows-1, end_column=cs+9)
            for i in range(rows):
                b_style = border_blue_bottom if i == rows - 1 else border_all_thin
                ws.cell(r+i, cs+9).border = b_style

            return r + rows

        # --- 資料分群與排序 (日紅沉底) ---
        rL, rR = 3, 3
        
        # 彰化區
        ch_d = df[df['區域']=='彰化']
        all_ch_stores = list(dict.fromkeys(ch_d['店名']))
        # 分離日紅與非日紅
        rihong_stores = [s for s in all_ch_stores if '日紅' in s]
        normal_stores = [s for s in all_ch_stores if '日紅' not in s]
        # 合併順序：先一般，後日紅
        sorted_ch_stores = normal_stores + rihong_stores
        
        for s in sorted_ch_stores: 
            rL = render_store(ch_d[ch_d['店名']==s], rL, 1)

        # 台中區 (直接排序)
        tc_d = df[df['區域']=='台中']
        for s in list(dict.fromkeys(tc_d['店名'])): 
            rR = render_store(tc_d[tc_d['店名']==s], rR, 12)

        # --- 底部統計 ---
        # 彰化底部
        ws.row_dimensions[rL].height = 22
        ws.cell(rL, 4, ch_d['檳榔'].sum()).font=font_green; ws.cell(rL, 7, ch_d['實收'].sum()).font=font_green
        for c in [4, 7]: ws.cell(rL, c).number_format='#,##0'; ws.cell(rL, c).alignment=align_c; ws.cell(rL, c).border=border_all_thin
        for c in range(1, 11): 
            if c not in [4, 7]: ws.cell(rL, c).border=border_all_thin
        rL += 1
        
        # 備註區 (調入調出)
        for lbl in ["班別入帳：", "轉入轉出：", "調入調出："]:
            ws.merge_cells(start_row=rL, start_column=1, end_row=rL+1, end_column=10)
            ws.cell(rL, 1, lbl).alignment=align_l_top; ws.cell(rL, 1, lbl).font=font_n # 字級14
            for ro in range(2): 
                ws.row_dimensions[rL+ro].height = 22 # 列高22
                for ci in range(1, 11): ws.cell(rL+ro, ci).border=border_all_thin
            rL += 2

        # 台中底部
        ws.row_dimensions[rR].height = 22
        ws.cell(rR, 15, tc_d['檳榔'].sum()).font=font_green; ws.cell(rR, 18, tc_d['實收'].sum()).font=font_green
        for c in [15, 18]: ws.cell(rR, c).number_format='#,##0'; ws.cell(rR, c).alignment=align_c; ws.cell(rR, c).border=border_all_thin
        for c in range(12, 22): 
            if c not in [15, 18]: ws.cell(rR, c).border=border_all_thin
        rR += 1
        
        # 今日全體大計
        ws.row_dimensions[rR].height = 22
        gr, gb, gd = df['實收'].sum(), ch_d['檳榔'].sum()+tc_d['檳榔'].sum(), df['帳差'].sum()
        ws.cell(rR, 15, gb).font=font_b; ws.cell(rR, 16, gr).font=font_b; ws.cell(rR, 18, gr).font=font_b
        cd = ws.cell(rR, 17, gd); cd.font=font_red if gd<0 else (font_blue if gd>0 else font_b)
        for c in [15, 16, 17, 18]: ws.cell(rR, c).number_format='#,##0'; ws.cell(rR, c).alignment=align_c
        for c in range(12, 22): ws.cell(rR, c).border=border_all_thin
        rR += 1

        # Panel B
        ms = report_date.replace(day=1)
        dr = f"{ms.month}/{ms.day}-{report_date.month}/{report_date.day}"
        pd_data = [
            (f"{dr} 營業總金額：", p_t + gr), 
            (f"{dr} 彰化檳榔金額：", p_ch + ch_d['檳榔'].sum()), 
            (f"{dr} 台中檳榔金額：", p_tc + tc_d['檳榔'].sum())
        ]
        curr = rR + 1
        for lbl, val in pd_data:
            ws.merge_cells(start_row=curr, start_column=12, end_row=curr+1, end_column=15)
            ws.cell(curr, 12, lbl).alignment=align_r; ws.cell(curr, 12, lbl).font=font_panel
            
            ws.merge_cells(start_row=curr, start_column=16, end_row=curr+1, end_column=19)
            ws.cell(curr, 16, val).number_format='#,##0'; ws.cell(curr, 16, val).font=font_panel; ws.cell(curr, 16, val).alignment=align_c
            
            for rr in range(curr, curr+2):
                ws.row_dimensions[rr].height = 22 # 列高22
                for cc in range(12, 20): ws.cell(rr, cc).border=border_all_thin
            curr += 2
            
        for lbl in ["彰化區未收款：", "台中區未收款：", "", "現金正負差：", "實收總金額："]:
            if lbl:
                ws.merge_cells(start_row=curr, start_column=12, end_row=curr+1, end_column=15)
                ws.cell(curr, 12, lbl).alignment=align_r; ws.cell(curr, 12, lbl).font=font_panel
                
                ws.merge_cells(start_row=curr, start_column=16, end_row=curr+1, end_column=19)
                for rr in range(curr, curr+2):
                    ws.row_dimensions[rr].height = 22 # 列高22
                    for cc in range(12, 20): ws.cell(rr, cc).border=border_all_thin
                curr += 2
            else: curr += 1

        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        out = io.BytesIO(); wb.save(out)
        st.success(f"✅ {sn} 報表生成完成！")
        st.download_button("💾 下載最新報表", out.getvalue(), f"日報表_{report_date.strftime('%Y%m%d')}.xlsx")
